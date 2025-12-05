"""
Accounting Assistant Service
會計助手服務

Features / 功能:
- Receipt upload and AI analysis / 收據上傳和AI分析
- Automatic categorization / 自動分類
- Auto double-entry bookkeeping / 自動複式記帳
- Multi-language receipt support / 多語言收據支援
- Excel report generation / Excel報表生成
- Excel comparison / Excel對比功能
- AI suggestions and error checking / AI建議和除錯
"""

import io
import json
import base64
import uuid
from datetime import datetime, date
from decimal import Decimal
from typing import Optional, List, Dict, Any
from django.conf import settings
from django.db import transaction
from openai import OpenAI

# Initialize OpenAI client
client = OpenAI(api_key=settings.OPENAI_API_KEY)


# =============================================================================
# Receipt Analysis Service / 收據分析服務
# =============================================================================

def analyze_receipt_image(image_base64: str, language: str = 'auto') -> Dict[str, Any]:
    """
    Analyze receipt image using GPT-4 Vision
    使用 GPT-4 Vision 分析收據圖片
    
    Args:
        image_base64: Base64 encoded image
        language: 'en', 'zh-TW', 'zh-CN', 'ja', 'auto' for auto-detect
    
    Returns:
        Extracted receipt data with AI analysis
    """
    
    language_instruction = {
        'auto': 'Detect the language of the receipt automatically.',
        'en': 'The receipt is in English.',
        'zh-TW': '收據是繁體中文。',
        'zh-CN': '收据是简体中文。',
        'ja': '領収書は日本語です。',
    }.get(language, 'Detect the language automatically.')
    
    prompt = f"""You are an expert accountant assistant. Analyze this receipt image and extract all relevant information.
{language_instruction}

Please extract and return a JSON object with the following structure:
{{
    "vendor_name": "Name of the store/vendor",
    "vendor_address": "Address if visible",
    "vendor_phone": "Phone number if visible",
    "vendor_tax_id": "Tax ID/VAT number if visible",
    "receipt_number": "Receipt/Invoice number",
    "receipt_date": "YYYY-MM-DD format",
    "receipt_time": "HH:MM format if visible",
    "currency": "Currency code (TWD, USD, JPY, etc.)",
    "items": [
        {{
            "description": "Item description",
            "quantity": 1,
            "unit_price": 0.00,
            "amount": 0.00,
            "tax_included": true
        }}
    ],
    "subtotal": 0.00,
    "tax_amount": 0.00,
    "tax_rate": 5,
    "discount_amount": 0.00,
    "total_amount": 0.00,
    "payment_method": "CASH/CREDIT_CARD/DEBIT_CARD/OTHER",
    "category_suggestion": "MEALS/TRANSPORTATION/OFFICE_SUPPLIES/UTILITIES/ENTERTAINMENT/OTHER",
    "expense_type": "OPERATING/COST_OF_GOODS/PAYROLL/RENT/UTILITIES/OTHER_EXPENSE",
    "detected_language": "en/zh-TW/zh-CN/ja/other",
    "confidence_score": 0.95,
    "notes": "Any additional notes or observations",
    "warnings": ["List any issues or unclear items"],
    "ai_suggestions": ["Suggestions for categorization or processing"]
}}

Important:
1. If any field is unclear or not visible, use null
2. Always try to extract the total amount
3. Suggest the most appropriate expense category
4. Note any anomalies or potential errors
5. Provide helpful suggestions for the accountant"""

    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/jpeg;base64,{image_base64}",
                                "detail": "high"
                            }
                        }
                    ]
                }
            ],
            max_tokens=2000,
            temperature=0.1
        )
        
        result_text = response.choices[0].message.content
        
        # Parse JSON from response
        try:
            # Try to extract JSON from the response
            if "```json" in result_text:
                json_str = result_text.split("```json")[1].split("```")[0]
            elif "```" in result_text:
                json_str = result_text.split("```")[1].split("```")[0]
            else:
                json_str = result_text
            
            result = json.loads(json_str.strip())
        except json.JSONDecodeError:
            result = {
                "error": "Failed to parse receipt",
                "raw_response": result_text,
                "confidence_score": 0.0
            }
        
        return result
        
    except Exception as e:
        return {
            "error": str(e),
            "confidence_score": 0.0
        }


def categorize_expense(receipt_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Categorize expense and suggest account mapping
    分類費用並建議會計科目對應
    """
    
    category_to_account = {
        'MEALS': {'code': '6300', 'name': '伙食費 / Meals', 'type': 'EXPENSE', 'subtype': 'OPERATING'},
        'TRANSPORTATION': {'code': '6200', 'name': '交通費 / Transportation', 'type': 'EXPENSE', 'subtype': 'OPERATING'},
        'OFFICE_SUPPLIES': {'code': '6100', 'name': '辦公用品 / Office Supplies', 'type': 'EXPENSE', 'subtype': 'OPERATING'},
        'UTILITIES': {'code': '6400', 'name': '水電費 / Utilities', 'type': 'EXPENSE', 'subtype': 'UTILITIES'},
        'ENTERTAINMENT': {'code': '6500', 'name': '交際費 / Entertainment', 'type': 'EXPENSE', 'subtype': 'OPERATING'},
        'RENT': {'code': '6600', 'name': '租金 / Rent', 'type': 'EXPENSE', 'subtype': 'RENT'},
        'TELEPHONE': {'code': '6700', 'name': '電話費 / Telephone', 'type': 'EXPENSE', 'subtype': 'UTILITIES'},
        'INSURANCE': {'code': '6800', 'name': '保險費 / Insurance', 'type': 'EXPENSE', 'subtype': 'OPERATING'},
        'MAINTENANCE': {'code': '6900', 'name': '維修費 / Maintenance', 'type': 'EXPENSE', 'subtype': 'OPERATING'},
        'OTHER': {'code': '6999', 'name': '其他費用 / Other Expenses', 'type': 'EXPENSE', 'subtype': 'OTHER_EXPENSE'},
    }
    
    category = receipt_data.get('category_suggestion', 'OTHER')
    account_info = category_to_account.get(category, category_to_account['OTHER'])
    
    return {
        'category': category,
        'suggested_account': account_info,
        'debit_account': account_info,  # Expense account
        'credit_account': {
            'code': '1100',
            'name': '銀行存款 / Bank',
            'type': 'ASSET',
            'subtype': 'BANK'
        } if receipt_data.get('payment_method') != 'CASH' else {
            'code': '1000',
            'name': '現金 / Cash',
            'type': 'ASSET',
            'subtype': 'CASH'
        }
    }


def generate_double_entry(receipt_data: Dict[str, Any], categorization: Dict[str, Any]) -> Dict[str, Any]:
    """
    Generate double-entry journal entry from receipt
    從收據生成複式記帳分錄
    """
    
    total_amount = Decimal(str(receipt_data.get('total_amount', 0)))
    tax_amount = Decimal(str(receipt_data.get('tax_amount', 0)))
    net_amount = total_amount - tax_amount
    
    journal_entry = {
        'entry_number': f"EXP-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}",
        'date': receipt_data.get('receipt_date', datetime.now().strftime('%Y-%m-%d')),
        'description': f"費用 - {receipt_data.get('vendor_name', 'Unknown Vendor')} / Expense - {receipt_data.get('vendor_name', 'Unknown Vendor')}",
        'reference': receipt_data.get('receipt_number', ''),
        'lines': [
            {
                'account_code': categorization['debit_account']['code'],
                'account_name': categorization['debit_account']['name'],
                'description': f"{receipt_data.get('vendor_name', '')} - {categorization['category']}",
                'debit': float(net_amount),
                'credit': 0,
            }
        ],
        'total_debit': float(total_amount),
        'total_credit': float(total_amount),
        'status': 'DRAFT',
        'ai_generated': True,
        'source_receipt': receipt_data.get('receipt_number', ''),
    }
    
    # Add tax line if applicable
    if tax_amount > 0:
        journal_entry['lines'].append({
            'account_code': '1150',
            'account_name': '進項稅額 / Input VAT',
            'description': f"VAT on {receipt_data.get('vendor_name', '')}",
            'debit': float(tax_amount),
            'credit': 0,
        })
    
    # Credit line (payment)
    journal_entry['lines'].append({
        'account_code': categorization['credit_account']['code'],
        'account_name': categorization['credit_account']['name'],
        'description': f"Payment to {receipt_data.get('vendor_name', '')}",
        'debit': 0,
        'credit': float(total_amount),
    })
    
    return journal_entry


# =============================================================================
# AI Suggestions Service / AI建議服務
# =============================================================================

def get_ai_suggestions(receipt_data: Dict[str, Any], journal_entry: Dict[str, Any]) -> Dict[str, Any]:
    """
    Get AI suggestions and error checking for the receipt and journal entry
    獲取AI建議和錯誤檢查
    """
    
    prompt = f"""You are an expert accountant. Review this receipt data and journal entry, then provide:
1. Validation of the categorization
2. Any potential errors or issues
3. Tax compliance suggestions
4. Cost optimization recommendations
5. Any anomalies detected

Receipt Data:
{json.dumps(receipt_data, indent=2, default=str)}

Generated Journal Entry:
{json.dumps(journal_entry, indent=2, default=str)}

Please respond in JSON format:
{{
    "validation_status": "VALID/WARNING/ERROR",
    "validation_issues": ["List of issues found"],
    "categorization_review": {{
        "is_correct": true/false,
        "suggested_category": "If different",
        "reason": "Explanation"
    }},
    "tax_compliance": {{
        "status": "OK/WARNING/ERROR",
        "notes": ["Tax-related observations"]
    }},
    "suggestions": [
        {{
            "type": "COST_SAVING/COMPLIANCE/PROCESS/OTHER",
            "title": "Suggestion title",
            "description": "Detailed suggestion",
            "priority": "HIGH/MEDIUM/LOW"
        }}
    ],
    "anomalies": ["List of detected anomalies"],
    "overall_score": 0.95,
    "summary": "Brief summary in both English and Chinese"
}}"""

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "system", "content": prompt}],
            max_tokens=1500,
            temperature=0.2
        )
        
        result_text = response.choices[0].message.content
        
        try:
            if "```json" in result_text:
                json_str = result_text.split("```json")[1].split("```")[0]
            elif "```" in result_text:
                json_str = result_text.split("```")[1].split("```")[0]
            else:
                json_str = result_text
            
            return json.loads(json_str.strip())
        except json.JSONDecodeError:
            return {
                "validation_status": "WARNING",
                "suggestions": [{"type": "OTHER", "title": "Manual Review Required", "description": result_text, "priority": "MEDIUM"}],
                "overall_score": 0.7
            }
            
    except Exception as e:
        return {
            "validation_status": "ERROR",
            "error": str(e),
            "overall_score": 0.0
        }


# =============================================================================
# Excel Report Generation / Excel報表生成
# =============================================================================

def generate_expense_report_excel(expenses: List[Dict[str, Any]], report_config: Dict[str, Any] = None) -> bytes:
    """
    Generate Excel expense report for manager approval
    生成費用報表Excel供主管簽核
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Report 費用報表"
    
    # Styles
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=16)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = f"費用報表 / Expense Report - {datetime.now().strftime('%Y-%m')}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Report info
    ws['A3'] = "報表日期 / Report Date:"
    ws['B3'] = datetime.now().strftime('%Y-%m-%d')
    ws['A4'] = "報表編號 / Report No:"
    ws['B4'] = f"EXP-RPT-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}"
    
    # Headers
    headers = [
        "日期\nDate", "收據編號\nReceipt No", "供應商\nVendor", 
        "類別\nCategory", "說明\nDescription", "金額\nAmount",
        "稅額\nTax", "總計\nTotal"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    ws.row_dimensions[6].height = 35
    
    # Data rows
    total_amount = Decimal('0')
    total_tax = Decimal('0')
    total_sum = Decimal('0')
    
    for row_idx, expense in enumerate(expenses, 7):
        ws.cell(row=row_idx, column=1, value=expense.get('date', '')).border = thin_border
        ws.cell(row=row_idx, column=2, value=expense.get('receipt_number', '')).border = thin_border
        ws.cell(row=row_idx, column=3, value=expense.get('vendor_name', '')).border = thin_border
        ws.cell(row=row_idx, column=4, value=expense.get('category', '')).border = thin_border
        ws.cell(row=row_idx, column=5, value=expense.get('description', '')).border = thin_border
        
        amount = Decimal(str(expense.get('subtotal', 0)))
        tax = Decimal(str(expense.get('tax_amount', 0)))
        total = Decimal(str(expense.get('total_amount', 0)))
        
        ws.cell(row=row_idx, column=6, value=float(amount)).border = thin_border
        ws.cell(row=row_idx, column=6).number_format = '#,##0.00'
        
        ws.cell(row=row_idx, column=7, value=float(tax)).border = thin_border
        ws.cell(row=row_idx, column=7).number_format = '#,##0.00'
        
        ws.cell(row=row_idx, column=8, value=float(total)).border = thin_border
        ws.cell(row=row_idx, column=8).number_format = '#,##0.00'
        
        total_amount += amount
        total_tax += tax
        total_sum += total
    
    # Total row
    total_row = len(expenses) + 7
    ws.cell(row=total_row, column=5, value="總計 / Total").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=float(total_amount)).font = Font(bold=True)
    ws.cell(row=total_row, column=6).number_format = '#,##0.00'
    ws.cell(row=total_row, column=7, value=float(total_tax)).font = Font(bold=True)
    ws.cell(row=total_row, column=7).number_format = '#,##0.00'
    ws.cell(row=total_row, column=8, value=float(total_sum)).font = Font(bold=True)
    ws.cell(row=total_row, column=8).number_format = '#,##0.00'
    
    # Signature section
    sig_row = total_row + 3
    ws.cell(row=sig_row, column=1, value="申請人 / Applicant:").font = Font(bold=True)
    ws.cell(row=sig_row, column=3, value="_________________")
    ws.cell(row=sig_row, column=5, value="日期 / Date:").font = Font(bold=True)
    ws.cell(row=sig_row, column=6, value="_________________")
    
    sig_row += 2
    ws.cell(row=sig_row, column=1, value="主管簽核 / Manager Approval:").font = Font(bold=True)
    ws.cell(row=sig_row, column=3, value="_________________")
    ws.cell(row=sig_row, column=5, value="日期 / Date:").font = Font(bold=True)
    ws.cell(row=sig_row, column=6, value="_________________")
    
    sig_row += 2
    ws.cell(row=sig_row, column=1, value="會計審核 / Accounting Review:").font = Font(bold=True)
    ws.cell(row=sig_row, column=3, value="_________________")
    ws.cell(row=sig_row, column=5, value="日期 / Date:").font = Font(bold=True)
    ws.cell(row=sig_row, column=6, value="_________________")
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# =============================================================================
# Excel Comparison Service / Excel對比服務
# =============================================================================

def compare_excel_with_database(excel_file: bytes, db_records: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Compare manually entered Excel data with database records
    比較手工Excel數據和資料庫記錄
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required")
    
    # Load Excel file
    wb = openpyxl.load_workbook(io.BytesIO(excel_file))
    ws = wb.active
    
    excel_records = []
    header_row = None
    
    # Find header row and extract data
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cells = [cell.value for cell in row]
        if any('date' in str(cell).lower() or '日期' in str(cell or '') for cell in cells):
            header_row = cells
            continue
        if header_row and any(cells):
            record = dict(zip(header_row, cells))
            if record.get(header_row[0]):  # Has data
                excel_records.append(record)
    
    # Compare records
    differences = {
        'missing_in_db': [],
        'missing_in_excel': [],
        'amount_mismatches': [],
        'data_mismatches': [],
        'matched': [],
    }
    
    # Create lookup for DB records
    db_lookup = {r.get('receipt_number', ''): r for r in db_records if r.get('receipt_number')}
    excel_lookup = {}
    
    for excel_rec in excel_records:
        # Try to find matching receipt number
        receipt_num = None
        for key in excel_rec:
            if key and ('receipt' in str(key).lower() or '收據' in str(key) or '編號' in str(key)):
                receipt_num = str(excel_rec[key]) if excel_rec[key] else None
                break
        
        if receipt_num:
            excel_lookup[receipt_num] = excel_rec
            
            if receipt_num in db_lookup:
                db_rec = db_lookup[receipt_num]
                
                # Compare amounts
                excel_amount = None
                db_amount = db_rec.get('total_amount', 0)
                
                for key in excel_rec:
                    if key and ('total' in str(key).lower() or '總計' in str(key) or '金額' in str(key)):
                        try:
                            excel_amount = float(excel_rec[key]) if excel_rec[key] else None
                        except (ValueError, TypeError):
                            pass
                        break
                
                if excel_amount is not None and abs(float(db_amount) - excel_amount) > 0.01:
                    differences['amount_mismatches'].append({
                        'receipt_number': receipt_num,
                        'excel_amount': excel_amount,
                        'db_amount': float(db_amount),
                        'difference': excel_amount - float(db_amount),
                    })
                else:
                    differences['matched'].append(receipt_num)
            else:
                differences['missing_in_db'].append({
                    'receipt_number': receipt_num,
                    'excel_data': excel_rec
                })
    
    # Find DB records not in Excel
    for receipt_num, db_rec in db_lookup.items():
        if receipt_num not in excel_lookup:
            differences['missing_in_excel'].append({
                'receipt_number': receipt_num,
                'db_data': db_rec
            })
    
    # Generate summary
    summary = {
        'total_excel_records': len(excel_records),
        'total_db_records': len(db_records),
        'matched_count': len(differences['matched']),
        'missing_in_db_count': len(differences['missing_in_db']),
        'missing_in_excel_count': len(differences['missing_in_excel']),
        'amount_mismatch_count': len(differences['amount_mismatches']),
        'differences': differences,
    }
    
    return summary


def get_comparison_ai_analysis(comparison_result: Dict[str, Any]) -> Dict[str, Any]:
    """
    Get AI analysis of the comparison results
    獲取比較結果的AI分析
    """
    
    prompt = f"""You are an expert auditor. Analyze this comparison between Excel data and database records.
Provide insights, identify potential issues, and suggest corrections.

Comparison Result:
{json.dumps(comparison_result, indent=2, default=str)}

Please respond in JSON format with both English and Chinese:
{{
    "analysis_summary": {{
        "en": "English summary",
        "zh": "中文摘要"
    }},
    "discrepancy_analysis": [
        {{
            "issue": "Description of issue",
            "severity": "HIGH/MEDIUM/LOW",
            "likely_cause": "Most probable reason",
            "recommendation": "Suggested action"
        }}
    ],
    "risk_assessment": {{
        "level": "LOW/MEDIUM/HIGH/CRITICAL",
        "factors": ["List of risk factors"]
    }},
    "action_items": [
        {{
            "priority": 1,
            "action": "What to do",
            "responsible": "Who should do it"
        }}
    ],
    "compliance_notes": ["Any compliance-related observations"],
    "overall_health_score": 0.85
}}"""

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "system", "content": prompt}],
            max_tokens=1500,
            temperature=0.2
        )
        
        result_text = response.choices[0].message.content
        
        try:
            if "```json" in result_text:
                json_str = result_text.split("```json")[1].split("```")[0]
            elif "```" in result_text:
                json_str = result_text.split("```")[1].split("```")[0]
            else:
                json_str = result_text
            
            return json.loads(json_str.strip())
        except json.JSONDecodeError:
            return {"analysis_summary": {"en": result_text, "zh": result_text}}
            
    except Exception as e:
        return {"error": str(e)}


# =============================================================================
# Full Receipt Processing Pipeline / 完整收據處理流程
# =============================================================================

def process_receipt_full(image_base64: str, language: str = 'auto', auto_save: bool = False) -> Dict[str, Any]:
    """
    Full receipt processing pipeline
    完整收據處理流程
    
    1. Analyze receipt image
    2. Categorize expense
    3. Generate double-entry
    4. Get AI suggestions
    5. Optionally save to database
    """
    
    result = {
        'status': 'processing',
        'steps': []
    }
    
    # Step 1: Analyze receipt
    result['steps'].append({'step': 'analyze', 'status': 'processing'})
    receipt_data = analyze_receipt_image(image_base64, language)
    
    if receipt_data.get('error'):
        result['status'] = 'error'
        result['error'] = receipt_data['error']
        result['steps'][-1]['status'] = 'error'
        return result
    
    result['receipt_data'] = receipt_data
    result['steps'][-1]['status'] = 'complete'
    
    # Step 2: Categorize
    result['steps'].append({'step': 'categorize', 'status': 'processing'})
    categorization = categorize_expense(receipt_data)
    result['categorization'] = categorization
    result['steps'][-1]['status'] = 'complete'
    
    # Step 3: Generate journal entry
    result['steps'].append({'step': 'journal_entry', 'status': 'processing'})
    journal_entry = generate_double_entry(receipt_data, categorization)
    result['journal_entry'] = journal_entry
    result['steps'][-1]['status'] = 'complete'
    
    # Step 4: AI suggestions
    result['steps'].append({'step': 'ai_review', 'status': 'processing'})
    ai_suggestions = get_ai_suggestions(receipt_data, journal_entry)
    result['ai_suggestions'] = ai_suggestions
    result['steps'][-1]['status'] = 'complete'
    
    # Step 5: Save to database (if requested)
    if auto_save:
        result['steps'].append({'step': 'save', 'status': 'processing'})
        # This would call the actual Django model save
        # For now, we just mark it as ready
        result['save_ready'] = True
        result['steps'][-1]['status'] = 'ready'
    
    result['status'] = 'complete'
    return result
